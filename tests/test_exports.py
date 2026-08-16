import io
import os
import zipfile

from datetime import date
from datetime import time

import pytest
from openpyxl import load_workbook

os.environ.setdefault('SECRET_KEY', 'test-secret-key')

import app as app_pkg
from app.constants import ROLE_ADMIN


MEMBERS = {
    'E001': {
        'name': 'Alice Smith',
        'department': 'Engineering',
        'staff_id': 'E001',
        'position': 'Developer',
        'level': 'Senior',
    },
    'E002': {
        'name': 'Bob / QA',
        'department': 'Quality',
        'staff_id': 'E002',
        'position': 'QA Engineer',
        'level': 'Mid',
    },
}

WORKLOGS = {
    'E001': [
        {
            'log_date': date(2026, 1, 5),
            'project': 'MWL',
            'task': 'Export tests',
            'projectdepartment': 'Platform',
            'description': 'Verify individual export',
            'start_time': time(8, 30),
            'end_time': time(17, 30),
            'hours': 8,
            'status': 'Done',
            'note': 'Validated',
        },
        {
            'log_date': date(2026, 1, 5),
            'project': 'MWL',
            'task': 'Workbook review',
            'projectdepartment': 'Platform',
            'description': 'Inspect formulas',
            'start_time': time(18, 0),
            'end_time': time(19, 30),
            'hours': 1.5,
            'status': 'In Progress',
            'note': '',
        },
    ],
    'E002': [
        {
            'log_date': '2026-01-06',
            'project': 'QA',
            'task': 'Bulk export',
            'projectdepartment': 'Quality',
            'description': 'Open generated workbook',
            'start_time': '09:00',
            'end_time': '12:00',
            'hours': 3,
            'status': 'Pending',
            'note': 'Follow up',
        },
    ],
}


@pytest.fixture()
def client(monkeypatch):
    def fake_query(sql, params=(), fetchone=False):
        if 'FROM Employee' in sql:
            member = MEMBERS.get(params[0])
            return dict(member) if member else None
        if 'FROM worklogs' in sql:
            return [dict(row) for row in WORKLOGS.get(params[0], [])]
        raise AssertionError(f'Unexpected export query: {sql}')

    monkeypatch.setattr(app_pkg.db, 'query', fake_query)
    app_pkg.app.config.update(TESTING=True)
    app_pkg.app._db_initialized = True

    with app_pkg.app.test_client() as test_client:
        with test_client.session_transaction() as user_session:
            user_session['user_id'] = 1
            user_session['role'] = ROLE_ADMIN
            user_session['member_id'] = 'E001'
        yield test_client


def _load_valid_workbook(raw_bytes):
    assert raw_bytes.startswith(b'PK')
    workbook = load_workbook(io.BytesIO(raw_bytes), data_only=False)
    assert workbook.sheetnames
    assert workbook.calculation.calcMode == 'auto'
    return workbook


def _assert_no_broken_formula_references(workbook):
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    assert '#REF!' not in cell.value
                    assert '[' not in cell.value


def test_individual_export_returns_valid_workbook(client):
    response = client.get('/api/export/excel?member_id=E001&year=2026&months=1,2')

    assert response.status_code == 200
    assert response.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert 'Alice_Smith_Worklog_2026.xlsx' in response.headers['Content-Disposition']

    workbook = _load_valid_workbook(response.data)
    assert workbook.sheetnames == ['Dashboard', 'January', 'February']
    assert workbook['Dashboard']['B2'].value == 'Alice Smith'
    assert workbook['Dashboard']['B3'].value == 'Engineering'
    assert workbook['January']['A5'].value == date(2026, 1, 5)
    assert workbook['January']['B5'].value == 'MWL'
    assert workbook['January']['H5'].value == 8
    assert workbook['January']['A6'].value is None
    assert workbook['January']['H6'].value == 1.5
    assert workbook['January']['M5'].value == '=SUM(H5:H6)'
    assert workbook['Dashboard']['B14'].value == "='January'!$M$5"
    assert workbook['Dashboard']['B15'].value == "='February'!$M$5"
    _assert_no_broken_formula_references(workbook)


def test_bulk_export_contains_valid_workbook_per_member(client):
    response = client.get('/api/export/excel/bulk?member_ids=E001,E002&year=2026&months=1')

    assert response.status_code == 200
    assert response.mimetype == 'application/zip'
    assert 'Team_Worklog_2026.zip' in response.headers['Content-Disposition']

    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        assert archive.testzip() is None
        assert sorted(archive.namelist()) == [
            'Alice_Smith_Worklog_2026.xlsx',
            'Bob___QA_Worklog_2026.xlsx',
        ]
        for filename, expected_name in (
            ('Alice_Smith_Worklog_2026.xlsx', 'Alice Smith'),
            ('Bob___QA_Worklog_2026.xlsx', 'Bob / QA'),
        ):
            workbook = _load_valid_workbook(archive.read(filename))
            assert workbook.sheetnames == ['Dashboard', 'January']
            assert workbook['Dashboard']['B2'].value == expected_name
            _assert_no_broken_formula_references(workbook)
