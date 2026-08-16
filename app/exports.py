import io
import re
import zipfile

from collections import defaultdict
from datetime import date
from datetime import datetime
from datetime import time

from flask import Blueprint
from flask import jsonify
from flask import request
from flask import send_file
from flask import session
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side

import app as app_pkg

from .auth import elevated_required
from .auth import login_required
from .constants import ELEVATED_ROLES
from .helpers import parse_time

export_bp = Blueprint('exports', __name__)

MONTH_NAMES = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December',
]

_DARK_BLUE = '2F5597'
_HEADER_BLUE = '4472C4'
_LIGHT_BLUE = 'D9EAF7'
_WHITE = 'FFFFFF'
_GRID = '7F8C8D'
_THIN_BORDER = Border(
    left=Side(style='thin', color=_GRID),
    right=Side(style='thin', color=_GRID),
    top=Side(style='thin', color=_GRID),
    bottom=Side(style='thin', color=_GRID),
)


def _safe_export_name(name):
    value = (name or '').strip()
    safe = re.sub(r'[^\w\-.]', '_', value)
    return safe or 'member'


def _normalise_months(months):
    if months is None:
        return list(range(1, 13))

    result = []
    for raw_month in months:
        month = int(raw_month)
        if month < 1 or month > 12:
            raise ValueError('months must contain values from 1 to 12')
        if month not in result:
            result.append(month)
    return result


def _set_text(cell, value):
    """Write user/database text as text, including values beginning with '='."""
    cell.value = '' if value is None else str(value)
    cell.data_type = 's'


def _excel_time(value):
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, str):
        return parse_time(value)
    return None


def _style_title(worksheet, cell_range):
    for row in worksheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor=_DARK_BLUE)
            cell.font = Font(color=_WHITE, bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')


def _style_header(worksheet, cell_range):
    for row in worksheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor=_HEADER_BLUE)
            cell.font = Font(color=_WHITE, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _THIN_BORDER


def _style_body(worksheet, cell_range):
    for row in worksheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor=_LIGHT_BLUE)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)


def _build_month_sheet(workbook, member, year, month_idx, row_data):
    sheet_name = MONTH_NAMES[month_idx - 1]
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = 'A5'

    worksheet.merge_cells('A1:J1')
    _set_text(worksheet['A1'], f'{sheet_name} {year} - Work Log')
    _style_title(worksheet, 'A1:J1')
    worksheet.row_dimensions[1].height = 24

    worksheet['A2'] = 'Name:'
    worksheet.merge_cells('B2:D2')
    _set_text(worksheet['B2'], member.get('name'))
    worksheet['E2'] = 'ID:'
    _set_text(worksheet['F2'], member.get('staff_id'))
    worksheet['G2'] = 'Position:'
    worksheet.merge_cells('H2:J2')
    _set_text(worksheet['H2'], member.get('position'))
    worksheet['A3'] = 'Department:'
    worksheet.merge_cells('B3:D3')
    _set_text(worksheet['B3'], member.get('department'))
    for cell in ('A2', 'E2', 'G2', 'A3'):
        worksheet[cell].font = Font(bold=True)

    headers = [
        'Date', 'Project code', 'Project department', 'Description', 'Task',
        'Start', 'End', 'Hours', 'Status', 'Note',
    ]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=4, column=column, value=header)
    _style_header(worksheet, 'A4:J4')
    worksheet.row_dimensions[4].height = 28

    last_row = max(5, 4 + len(row_data))
    _style_body(worksheet, f'A5:J{last_row}')
    for index, (entry_date, entry) in enumerate(row_data, start=5):
        if entry_date is not None:
            worksheet.cell(row=index, column=1, value=entry_date)
            worksheet.cell(row=index, column=1).number_format = 'dd/mm/yyyy'

        _set_text(worksheet.cell(row=index, column=2), entry.get('project'))
        _set_text(worksheet.cell(row=index, column=3), entry.get('projectdepartment'))
        _set_text(worksheet.cell(row=index, column=4), entry.get('description'))
        _set_text(worksheet.cell(row=index, column=5), entry.get('task'))

        start_time = _excel_time(entry.get('start_time'))
        end_time = _excel_time(entry.get('end_time'))
        if start_time is not None:
            worksheet.cell(row=index, column=6, value=start_time).number_format = 'h:mm'
        if end_time is not None:
            worksheet.cell(row=index, column=7, value=end_time).number_format = 'h:mm'
        if entry.get('hours') is not None:
            worksheet.cell(row=index, column=8, value=float(entry['hours'])).number_format = '0.00'
        _set_text(worksheet.cell(row=index, column=9), entry.get('status'))
        _set_text(worksheet.cell(row=index, column=10), entry.get('note'))

    worksheet.merge_cells('L4:M4')
    worksheet['L4'] = 'MONTHLY SUMMARY'
    _style_header(worksheet, 'L4:M4')
    summary_labels = ['Total Hours', 'Tasks Done', 'Tasks In Progress', 'Tasks Pending']
    for row_number, label in enumerate(summary_labels, start=5):
        worksheet.cell(row=row_number, column=12, value=label)
        worksheet.cell(row=row_number, column=12).font = Font(bold=True)
        worksheet.cell(row=row_number, column=12).fill = PatternFill('solid', fgColor=_LIGHT_BLUE)
        worksheet.cell(row=row_number, column=12).border = _THIN_BORDER
        worksheet.cell(row=row_number, column=13).border = _THIN_BORDER
    worksheet['M5'] = f'=SUM(H5:H{last_row})'
    worksheet['M6'] = f'=COUNTIF(I5:I{last_row},"Done")'
    worksheet['M7'] = f'=COUNTIF(I5:I{last_row},"In Progress")'
    worksheet['M8'] = f'=COUNTIF(I5:I{last_row},"Pending")'
    worksheet['M5'].number_format = '0.00'
    for cell in ('M5', 'M6', 'M7', 'M8'):
        worksheet[cell].alignment = Alignment(horizontal='right')

    widths = {
        'A': 13, 'B': 17, 'C': 21, 'D': 24, 'E': 32,
        'F': 10, 'G': 10, 'H': 10, 'I': 17, 'J': 24,
        'K': 3, 'L': 22, 'M': 13,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.auto_filter.ref = f'A4:J{last_row}'
    worksheet.print_title_rows = '1:4'
    worksheet.print_area = f'A1:M{last_row}'
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    return worksheet


def _build_dashboard(workbook, member, year, generated_months):
    dashboard = workbook.create_sheet('Dashboard', 0)
    dashboard.sheet_view.showGridLines = False
    dashboard.freeze_panes = 'A14'

    dashboard.merge_cells('A1:E1')
    _set_text(dashboard['A1'], f'{year} ANNUAL WORKLOG DASHBOARD')
    _style_title(dashboard, 'A1:E1')
    dashboard.row_dimensions[1].height = 26

    dashboard['A2'] = 'Name:'
    dashboard.merge_cells('B2:C2')
    _set_text(dashboard['B2'], member.get('name'))
    dashboard['D2'] = 'ID:'
    _set_text(dashboard['E2'], member.get('staff_id'))
    dashboard['A3'] = 'Department:'
    dashboard.merge_cells('B3:C3')
    _set_text(dashboard['B3'], member.get('department'))
    dashboard['D3'] = 'Position:'
    _set_text(dashboard['E3'], member.get('position'))
    for cell in ('A2', 'D2', 'A3', 'D3'):
        dashboard[cell].font = Font(bold=True)

    dashboard.merge_cells('A5:B5')
    dashboard['A5'] = 'ANNUAL SUMMARY'
    _style_header(dashboard, 'A5:B5')
    summary = [
        ('Total Hours (All Months)', '=SUM(B14:B25)', '0.00'),
        ('Total Tasks Done', '=SUM(C14:C25)', '0'),
        ('Total Tasks In Progress', '=SUM(D14:D25)', '0'),
        ('Total Tasks Pending', '=SUM(E14:E25)', '0'),
        ('Average Monthly Hours', '=ROUND(B6/12,2)', '0.00'),
    ]
    for row_number, (label, formula, number_format) in enumerate(summary, start=6):
        dashboard.cell(row=row_number, column=1, value=label)
        dashboard.cell(row=row_number, column=2, value=formula)
        dashboard.cell(row=row_number, column=2).number_format = number_format
    _style_body(dashboard, 'A6:B10')
    for row_number in range(6, 11):
        dashboard.cell(row=row_number, column=1).font = Font(bold=True)

    dashboard.merge_cells('A12:E12')
    dashboard['A12'] = 'MONTHLY BREAKDOWN'
    _style_header(dashboard, 'A12:E12')
    breakdown_headers = ['Month', 'Total Hours', 'Tasks Done', 'Tasks In Progress', 'Tasks Pending']
    for column, header in enumerate(breakdown_headers, start=1):
        dashboard.cell(row=13, column=column, value=header)
    _style_header(dashboard, 'A13:E13')

    generated_months = set(generated_months)
    for month_idx, sheet_name in enumerate(MONTH_NAMES, start=1):
        row_number = 13 + month_idx
        dashboard.cell(row=row_number, column=1, value=sheet_name)
        if month_idx in generated_months:
            dashboard.cell(row=row_number, column=2, value=f"='{sheet_name}'!$M$5")
            dashboard.cell(row=row_number, column=3, value=f"='{sheet_name}'!$M$6")
            dashboard.cell(row=row_number, column=4, value=f"='{sheet_name}'!$M$7")
            dashboard.cell(row=row_number, column=5, value=f"='{sheet_name}'!$M$8")
        else:
            for column in range(2, 6):
                dashboard.cell(row=row_number, column=column, value=0)
        dashboard.cell(row=row_number, column=2).number_format = '0.00'
    _style_body(dashboard, 'A14:E25')

    for column, width in {'A': 29, 'B': 18, 'C': 18, 'D': 22, 'E': 18}.items():
        dashboard.column_dimensions[column].width = width
    dashboard.print_area = 'A1:E25'
    dashboard.page_setup.fitToWidth = 1
    dashboard.sheet_properties.pageSetUpPr.fitToPage = True
    return dashboard


def generate_excel_bytes(employee_id, member, year, months=None):
    """Generate a filled Excel workbook for one employee and return raw bytes.

    `employee_id` is dbo.Employee.EmployeeID (post-migration); param name is
    kept generic to avoid changing callers.
    """
    first_day = date(year, 1, 1)
    last_day = date(year, 12, 31)
    worklogs = app_pkg.db.query(
        """
        SELECT log_date, project, task, projectdepartment, description,
               start_time, end_time, hours, status, note
        FROM worklogs
        WHERE EmployeeID = ? AND log_date BETWEEN ? AND ?
        ORDER BY log_date, start_time
        """,
        (employee_id, first_day, last_day),
    )

    months_to_generate = _normalise_months(months)
    by_month_day = defaultdict(list)
    for worklog in worklogs:
        log_date = worklog['log_date']
        if isinstance(log_date, str):
            log_date = datetime.strptime(log_date, '%Y-%m-%d').date()
        by_month_day[(log_date.month, log_date.day)].append(worklog)

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.iso_dates = True
    workbook.properties.creator = 'Meter Worklog'
    workbook.properties.title = f'{member.get("name") or "Member"} Worklog {year}'
    workbook.calculation.calcMode = 'auto'
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    for month_idx in months_to_generate:
        row_data = []
        month_days = sorted(
            day for (entry_month, day) in by_month_day
            if entry_month == month_idx
        )
        for day in month_days:
            current_date = date(year, month_idx, day)
            entries = by_month_day.get((month_idx, day), [])
            for index, entry in enumerate(entries):
                row_data.append((current_date if index == 0 else None, entry))
        _build_month_sheet(workbook, member, year, month_idx, row_data)

    _build_dashboard(workbook, member, year, months_to_generate)
    workbook.active = 0

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@export_bp.route('/api/export/excel', methods=['GET'])
@login_required
def export_excel():
    member_id = (request.args.get('member_id') or '').strip()
    year = request.args.get('year', type=int, default=date.today().year)
    months_str = request.args.get('months', '')
    months = [int(month) for month in months_str.split(',') if month.strip().isdigit()] or None

    if months and any(month < 1 or month > 12 for month in months):
        return jsonify({'error': 'months must contain values from 1 to 12'}), 400

    if not member_id:
        return jsonify({'error': 'member_id required'}), 400

    if not app_pkg._worklog_open and session.get('role') not in ELEVATED_ROLES and member_id != session.get('member_id'):
        return jsonify({'error': 'Permission denied'}), 403

    member = app_pkg.db.query(
        """SELECT EmployeeName AS name,
                  Department  AS department,
                  EmployeeID  AS staff_id,
                  Position    AS position,
                  Level       AS level
           FROM Employee WHERE EmployeeID=?""",
        (member_id,),
        fetchone=True,
    )
    if not member:
        return jsonify({'error': 'member not found'}), 404

    excel_bytes = generate_excel_bytes(member_id, member, year, months)
    safe_name = _safe_export_name(member['name'])
    return send_file(
        io.BytesIO(excel_bytes),
        as_attachment=True,
        download_name=f'{safe_name}_Worklog_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@export_bp.route('/api/export/excel/bulk', methods=['GET'])
@elevated_required
def export_excel_bulk():
    member_ids_str = request.args.get('member_ids', '')
    year = request.args.get('year', type=int, default=date.today().year)
    months_str = request.args.get('months', '')
    months = [int(month) for month in months_str.split(',') if month.strip().isdigit()] or None

    if months and any(month < 1 or month > 12 for month in months):
        return jsonify({'error': 'months must contain values from 1 to 12'}), 400

    if not member_ids_str:
        return jsonify({'error': 'member_ids required'}), 400

    member_ids = [m.strip() for m in member_ids_str.split(',') if m.strip()]

    if not member_ids:
        return jsonify({'error': 'No valid member_ids provided'}), 400

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
        for member_id in member_ids:
            member = app_pkg.db.query(
                """SELECT EmployeeName AS name,
                          Department  AS department,
                          EmployeeID  AS staff_id,
                          Position    AS position,
                          Level       AS level
                   FROM Employee WHERE EmployeeID=?""",
                (member_id,),
                fetchone=True,
            )
            if not member:
                continue

            excel_bytes = generate_excel_bytes(member_id, member, year, months)
            safe_name = _safe_export_name(member['name'])
            archive.writestr(f'{safe_name}_Worklog_{year}.xlsx', excel_bytes)

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=f'Team_Worklog_{year}.zip',
        mimetype='application/zip',
    )
